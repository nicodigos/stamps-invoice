from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path(__file__).resolve().parents[1] / "Cuentas Bancarias NO MOVER.xlsx"

COMPANIES = [
    "1001298527 ONTARIO INC",
    "10342548 CANADA INC",
    "10377180 CANADA INC-standby",
    "10696480 CANADA LTD",
    "12433087 CANADA INC-MASTER",
    "13037622 CANADA INC",
    "14661796 CANADA LTD- not active",
    "16021166 Canada Inc-Diego-Chacho",
    "9359-6633 QUEBEC INC",
    "9390-9216 QUEBEC INC",
    "D-TECH CONSTRUCTION",
    "TAYANTI-CANADA",
]

BANKS = ["Desjardin", "National", "Scotiabank"]
TYPES = ["debit", "credit"]

ROWS = [
    ["TAYANTI-CANADA", "Desjardin", "debit", "N0931"],
    ["TAYANTI-CANADA", "Desjardin", "credit", "N9014"],
    ["TAYANTI-CANADA", "Desjardin", "credit", "N0034"],
    ["12433087 CANADA INC-MASTER", "Desjardin", "credit", "N2024"],
    ["12433087 CANADA INC-MASTER", "Desjardin", "credit", "N1018"],
    ["10696480 CANADA LTD", "National", "credit", "N1380"],
    ["9359-6633 QUEBEC INC", "National", "debit", "N6413"],
    ["9359-6633 QUEBEC INC", "Desjardin", "credit", "N4013"],
]


def quote_sheet_range(sheet_name, start, end):
    return f"'{sheet_name}'!${start}:${end}"


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuentas Bancarias"

    lists = wb.create_sheet("_validations")
    lists.sheet_state = "hidden"

    for row_idx, company in enumerate(COMPANIES, start=1):
        lists.cell(row=row_idx, column=1, value=company)
    for row_idx, bank in enumerate(BANKS, start=1):
        lists.cell(row=row_idx, column=2, value=bank)
    for row_idx, account_type in enumerate(TYPES, start=1):
        lists.cell(row=row_idx, column=3, value=account_type)

    headers = ["Company", "Bank", "Type", "Last 4 digits"]
    ws.append(headers)
    for row in ROWS:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="DCEBF3")
    thin = Side(style="thin", color="D7E1E9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="203040")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left")
            cell.number_format = "@"

    widths = {
        "A": 34,
        "B": 18,
        "C": 12,
        "D": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    company_dv = DataValidation(
        type="list",
        formula1=quote_sheet_range("_validations", "A$1", f"A${len(COMPANIES)}"),
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid company",
        error="Choose a company from the allowed list.",
    )
    bank_dv = DataValidation(
        type="list",
        formula1=quote_sheet_range("_validations", "B$1", f"B${len(BANKS)}"),
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid bank",
        error="Choose a bank from the allowed list.",
    )
    type_dv = DataValidation(
        type="list",
        formula1=quote_sheet_range("_validations", "C$1", f"C${len(TYPES)}"),
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid type",
        error="Choose debit or credit.",
    )
    digits_dv = DataValidation(
        type="custom",
        formula1='=AND(LEN(D2)=5,LEFT(D2,1)="N",ISNUMBER(VALUE(RIGHT(D2,4))))',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid digits",
        error="Use N followed by exactly four digits, for example N0931.",
    )

    for dv, cells in (
        (company_dv, "A2:A1000"),
        (bank_dv, "B2:B1000"),
        (type_dv, "C2:C1000"),
        (digits_dv, "D2:D1000"),
    ):
        ws.add_data_validation(dv)
        dv.add(cells)

    return wb


def main():
    wb = build_workbook()
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
